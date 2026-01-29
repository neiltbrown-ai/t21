#!/usr/bin/env python3
"""
T21 Directory - Supabase Data Import Script

Usage:
  1. Install dependencies: pip install requests openpyxl
  2. Place your Excel files in the same directory (or update paths below)
  3. Run: python supabase_import.py

This script imports data from your Excel files into Supabase.
"""

import requests
import openpyxl
from datetime import datetime
import os

# ============== CONFIGURATION ==============
SUPABASE_URL = "https://qistidaxuevycutiegsa.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InFpc3RpZGF4dWV2eWN1dGllZ3NhIiwicm9sZSI6ImFub24iLCJpYXQiOjE3Njk3MTM5NTAsImV4cCI6MjA4NTI4OTk1MH0.6U6g4gsabRGxvcPAaO1so5cZgS38GqGhKfHmq6E9dSA"

# Update these paths to match your file locations
FINANCIAL_FILE = "down_syndrome_financial_resources_database.xlsx"
THERAPY_FILE = "ds_therapy_services_verified.xlsx"
INSPIRATION_FILE = "ds_inspiration_database.xlsx"

# ============== HELPERS ==============
headers = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=minimal"
}

def clean_value(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.isoformat()
    return str(val).strip() if val else None

def clean_numeric(val):
    if val is None:
        return None
    try:
        return float(val)
    except:
        return None

def clean_int(val):
    if val is None:
        return None
    try:
        return int(val)
    except:
        return None

def clean_bool(val):
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    return str(val).lower() in ('yes', 'true', '1', 'y')

def insert_records(table_name, records):
    """Insert records into Supabase table"""
    if not records:
        print(f"  No records to insert for {table_name}")
        return

    resp = requests.post(
        f"{SUPABASE_URL}/rest/v1/{table_name}",
        headers=headers,
        json=records
    )
    if resp.status_code in (200, 201):
        print(f"  ✓ Inserted {len(records)} records into {table_name}")
    else:
        print(f"  ✗ Error inserting into {table_name}: {resp.status_code}")
        print(f"    {resp.text[:500]}")

# ============== FINANCIAL RESOURCES ==============
def import_financial():
    if not os.path.exists(FINANCIAL_FILE):
        print(f"  ✗ File not found: {FINANCIAL_FILE}")
        return

    print("Loading Financial Resources...")
    wb = openpyxl.load_workbook(FINANCIAL_FILE, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    records = []
    for row in rows:
        if not row[1]:  # Skip if no program name
            continue
        record = {
            "program_id": clean_value(row[0]),
            "program_name": clean_value(row[1]),
            "organization_type": clean_value(row[2]),
            "website": clean_value(row[3]),
            "phone": clean_value(row[4]),
            "email": clean_value(row[5]),
            "address": clean_value(row[6]),
            "geographic_coverage": clean_value(row[7]),
            "states_available": clean_value(row[8]),
            "program_category": clean_value(row[9]),
            "assistance_type": clean_value(row[10]),
            "award_amount_min": clean_numeric(row[11]),
            "award_amount_max": clean_numeric(row[12]),
            "annual_cap": clean_numeric(row[13]),
            "lifetime_cap": clean_numeric(row[14]),
            "age_range_min": clean_int(row[15]) or 0,
            "age_range_max": clean_int(row[16]) or 999,
            "diagnosis_required": clean_value(row[17]),
            "income_limit": clean_value(row[18]),
            "income_limit_details": clean_value(row[19]),
            "asset_limit": clean_value(row[20]),
            "other_eligibility": clean_value(row[21]),
            "covered_expenses": clean_value(row[22]),
            "application_deadline": clean_value(row[23]),
            "application_type": clean_value(row[24]),
            "application_url": clean_value(row[25]),
            "reapplication_allowed": clean_value(row[26]),
            "processing_time": clean_value(row[27]),
            "program_description": clean_value(row[28]),
            "application_process": clean_value(row[29]),
            "key_features": clean_value(row[30]),
            "real_world_context": clean_value(row[31]),
            "special_notes": clean_value(row[32]) if len(row) > 32 else None,
        }
        records.append(record)

    print(f"  Parsed {len(records)} financial records")
    insert_records("financial_resources", records)

# ============== THERAPY SERVICES ==============
def import_therapy():
    if not os.path.exists(THERAPY_FILE):
        print(f"  ✗ File not found: {THERAPY_FILE}")
        return

    print("\nLoading Therapy Services...")
    wb = openpyxl.load_workbook(THERAPY_FILE, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    records = []
    for row in rows:
        if not row[1]:
            continue
        record = {
            "resource_id": clean_value(row[0]),
            "resource_name": clean_value(row[1]),
            "organization_name": clean_value(row[2]),
            "organization_type": clean_value(row[3]),
            "primary_category": clean_value(row[4]) or "Healthcare & Therapy",
            "subcategories": clean_value(row[5]),
            "resource_type": clean_value(row[6]),
            "ds_specificity": clean_value(row[7]),
            "website": clean_value(row[8]),
            "phone": clean_value(row[9]),
            "email": clean_value(row[10]),
            "address": clean_value(row[11]),
            "jurisdiction_level": clean_value(row[12]),
            "states_available": clean_value(row[13]),
            "service_area_notes": clean_value(row[14]),
            "lifecycle_stages": clean_value(row[15]),
            "age_min": clean_int(row[16]) or 0,
            "age_max": clean_int(row[17]) or 999,
            "eligibility_criteria": clean_value(row[18]),
            "cost_type": clean_value(row[19]),
            "cost_details": clean_value(row[20]),
            "application_status": clean_value(row[21]),
            "short_description": clean_value(row[22]),
            "full_description": clean_value(row[23]),
            "key_features": clean_value(row[24]),
            "practical_notes": clean_value(row[25]),
            "tags": clean_value(row[30]) if len(row) > 30 else None,
        }
        records.append(record)

    print(f"  Parsed {len(records)} therapy records")
    insert_records("therapy_services", records)

# ============== INSPIRATION PROFILES ==============
def import_inspiration():
    if not os.path.exists(INSPIRATION_FILE):
        print(f"  ✗ File not found: {INSPIRATION_FILE}")
        return

    print("\nLoading Inspiration Profiles...")
    wb = openpyxl.load_workbook(INSPIRATION_FILE, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    records = []
    for row in rows:
        if not row[1]:
            continue
        record = {
            "profile_id": clean_value(row[0]),
            "full_name": clean_value(row[1]),
            "known_as": clean_value(row[2]),
            "birth_year": clean_int(row[3]),
            "location_city": clean_value(row[4]),
            "location_state": clean_value(row[5]),
            "location_country": clean_value(row[6]) or "United States",
            "primary_field": clean_value(row[7]),
            "secondary_fields": clean_value(row[8]),
            "specific_achievements": clean_value(row[9]),
            "active_status": clean_value(row[10]),
            "active_since": clean_value(row[11]),
            "website": clean_value(row[12]),
            "instagram": clean_value(row[13]),
            "tiktok": clean_value(row[14]),
            "youtube": clean_value(row[15]),
            "facebook": clean_value(row[16]),
            "short_bio": clean_value(row[17]),
            "notable_quotes": clean_value(row[18]),
            "key_accomplishments": clean_value(row[19]),
            "speaking_available": clean_bool(row[20]) if len(row) > 20 else False,
            "awards_honors": clean_value(row[21]) if len(row) > 21 else None,
            "include_in_directory": clean_bool(row[22]) if len(row) > 22 else True,
            "directory_categories": clean_value(row[23]) if len(row) > 23 else None,
            "featured_profile": clean_bool(row[24]) if len(row) > 24 else False,
        }
        records.append(record)

    print(f"  Parsed {len(records)} inspiration records")
    insert_records("inspiration_profiles", records)

# ============== MAIN ==============
if __name__ == "__main__":
    print("="*50)
    print("T21 DIRECTORY - SUPABASE DATA IMPORT")
    print("="*50 + "\n")

    import_financial()
    import_therapy()
    import_inspiration()

    print("\n" + "="*50)
    print("IMPORT COMPLETE")
    print("="*50)
    print("\nCheck your Supabase Table Editor to verify the data.")
